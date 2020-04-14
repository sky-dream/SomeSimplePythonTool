# -*- coding: utf-8 -*-
"""
Created on Wed Apr 10 09:34:09 2019

"""
import re
import xlwt
from openpyxl import  Workbook
from openpyxl  import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import log

logger = log.setup_logger()

writestyle0 =  xlwt.easyxf('font: bold False; align: wrap on; borders: left THIN, right THIN, top THIN, bottom THIN; pattern: pattern solid, fore_colour white;')
writestyle1 =  xlwt.easyxf('font: bold True;  align: wrap on; borders: left THIN, right THIN, top THIN, bottom THIN; pattern: pattern solid, fore_colour yellow;')
writestyle2 =  xlwt.easyxf('font: bold True;  align: wrap on; borders: left THIN, right THIN, top THIN, bottom THIN; pattern: pattern solid, fore_colour gold;')
writestyle3 =  xlwt.easyxf('font: bold True;  align: wrap on; borders: left THIN, right THIN, top THIN, bottom THIN; pattern: pattern solid, fore_colour bright_green;')

Excel_header = [
            #mandatory part
            "CAN Frame ID(NET-PR input)(mandatory)", #A 0
            "CANMatrix  Signals(NET-PR input)(mandatory)",
            "Signal Owner(PCM input)(mandatory)",
            "Realized in BSS(FO input)(mandatory)（required in which BSS or rejected）",
            "ASW interface signal(FO input)(mandatory)",
            "Mapping relationship(FO input)(mandatory)",
            "FW based on varcode(DSW need FO input,need monitoring os not)(mandatory)",
            "FW(DSW need NET-PR input)(mandatory)",
            "DSW Node & Need export node status to ASW or not(FO input)(mandatory)",
            "DSW STM settings(FO input)(mandatory)", #J 9
            #ASW part
            "ASW internal signal name(FO input)(optional)", #K10
            "General Description:(FO input)(optional)",
            "Signal Type(FO input)(optional)",
            "Data Size(FO input)(optional)",
            "Phy Min(FO input)(optional)",
            "Phy Max(FO input)(optional)",
            "Initial Value(FO input)(optional)",
            "Meaning(FO input)(optional)",
            "Period(FO input)(optional)",
            "Unit(FO input)(optional)",
            "Resolution(FO input)(optional)", #U20
            #NET part
            "CAN Factor(NET-PR input)(optional)", #V21
            "CAN Offset(NET-PR input)(optional)",
            "CAN Min(NET-PR input)(optional)",
            "CAN Max(NET-PR input)(optional)",
            "CAN value description(NET-PR input)(optional)",
            "Comment(optional)"#AA 26
            ]

Col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
       'AA','AB', 'AC' ]

Total_Col = len(Excel_header)#27
Col_ASW_n = 10
Col_NET_n = 21
pre_CMXinfoMessage_Name = ""

RGB_Black = '000000'
RGB_White = 'FFFFFF'
RGB_Red = 'FF0000'
RGB_Yellow = 'FFFF00'
RGB_Orange = 'FFB90F'
RGB_Green = '76EE00'

font1 = Font(name='Arial',size=10,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='000000')
font2 = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='000000')
border1 = Border(left=Side(style='thin',color='000000'),
                 right=Side(style='thin',color='000000'),
                 top=Side(style='thin',color='000000'),
                 bottom=Side(style='thin',color='000000'),
                 diagonal=Side(style='thin',color='000000'),
                 diagonal_direction=0,
                 outline=Side(style='thin',color='000000'),
                 vertical=Side(style='thin',color='000000'),
                 horizontal=Side(style='thin',color='000000'))
align1 = Alignment(horizontal='left',vertical='center',wrap_text=True)
fill1 = PatternFill(start_color = RGB_Yellow, end_color = RGB_Yellow, fill_type = 'solid')
fill2 = PatternFill(start_color = RGB_Orange, end_color = RGB_Orange, fill_type = 'solid')
fill3 = PatternFill(start_color = RGB_Green, end_color = RGB_Green, fill_type = 'solid')
#workbook1 = Workbook()
#workbook1.add_named_style(style1)

def ExcelWrite1(filepath):
    #load exit Excle
    workbook = load_workbook(filepath)

    Function_Mapping_sheet = workbook.create_sheet("Function_Mapping", 0)

    for i in range(0, Total_Col):#write sheet header
        Function_Mapping_sheet.column_dimensions[Col[i]].width = 30
        if i <= 9:
            Function_Mapping_sheet.cell(row = 1, column = i+1, value = Excel_header[i])
            Function_Mapping_sheet.cell(row = 1, column = i+1).fill = fill1
            Function_Mapping_sheet.cell(row = 1, column = i+1).font = font1
        elif i > 9 and i <= 20:
            Function_Mapping_sheet.cell(row = 1, column = i+1, value = Excel_header[i])
            Function_Mapping_sheet.cell(row = 1, column = i+1).fill = fill2
            Function_Mapping_sheet.cell(row = 1, column = i+1).font = font1
        elif i > 20:
            Function_Mapping_sheet.cell(row = 1, column = i+1, value = Excel_header[i])
            Function_Mapping_sheet.cell(row = 1, column = i+1).fill = fill3
            Function_Mapping_sheet.cell(row = 1, column = i+1).font = font1
    return(workbook, Function_Mapping_sheet)


def Write_NETInfo1(sheetwrite, n, CMXinfo):
    global pre_CMXinfoMessage_Name

    for i in range(0, Total_Col):
        sheetwrite.cell(row = n+1, column = i+1, value = "")

    if pre_CMXinfoMessage_Name != CMXinfo.Message_Name:

        sheetwrite.cell(row = n+1, column = 1, value = CMXinfo.Message_Name) #write message row
        sheetwrite.cell(row = n+1, column = 2, value = "CAN_Message: " + CMXinfo.Message_Name +
                                                        "\nID: " + CMXinfo. Message_ID_Hex +
                                                        "\nDLC: " + CMXinfo. Message_Length +
                                                        "\nCycle Time: " + CMXinfo. Message_Cycletime +
                                                        "\nmessage related monitoring") #write message row
        sheetwrite.cell(row = n+1, column = 1).font = font2
        sheetwrite.cell(row = n+1, column = 2).font = font2

        pre_CMXinfoMessage_Name = CMXinfo.Message_Name
        n = n + 1
        for i in range(0, Total_Col):
            sheetwrite.cell(row = n+1, column = i+1, value = "")

    #NET part
    sheetwrite.cell(row = n+1, column = 1, value = CMXinfo.Message_Name)
    sheetwrite.cell(row = n+1, column = 2, value = CMXinfo.Signal_Name)

    i1 = Col_NET_n
    sheetwrite.cell(row = n+1, column = i1+1, value = CMXinfo.Signal_Factor); i1 += 1
    sheetwrite.cell(row = n+1, column = i1+1, value = CMXinfo.Signal_Offset); i1 += 1
    sheetwrite.cell(row = n+1, column = i1+1, value = CMXinfo.Signal_PhyMin); i1 += 1
    sheetwrite.cell(row = n+1, column = i1+1, value = CMXinfo.Signal_PhyMax); i1 += 1
    sheetwrite.cell(row = n+1, column = i1+1, value = CMXinfo.Signal_ValueTab); i1 += 1
    sheetwrite.cell(row = n+1, column = i1+1, value = 'InvalidVal: ' + CMXinfo.Signal_InvalidVal)

    #ASW part
    i2 = Col_ASW_n
    ASWDatabase_sheetname = "ASWDatabase"
    ASWDatabase_index = ['2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']

    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))");  i2 += 1
    sheetwrite.cell(row = n+1, column = i2+1, value = "=IF(ISNA(VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE)),\"\"," + "VLOOKUP(E$1:E$65536," + str(ASWDatabase_sheetname) + "!A$1:Z$65536," + ASWDatabase_index[i2-10] + ",FALSE))")

    return n

def Adjust_Style(sheetwrite):

    rows = sheetwrite.max_row
    cols = sheetwrite.max_column

    for i in range(1, rows+1):
        for j in range(1, cols+1):
            sheetwrite.cell(row = i, column = j).border = border1
            sheetwrite.cell(row = i, column = j).alignment = align1

def ExcelWrite():
    # 创建一个Workbook对象，这就相当于创建了一个Excel文件
    bookwrite = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheetwrite = bookwrite.add_sheet('Sheet1', cell_overwrite_ok=True)
    sheetwrite2 = bookwrite.add_sheet('Sheet2', cell_overwrite_ok=True)

    for i in range(0, Total_Col):#write sheet header
        sheetwrite.col(i).width = 256*30
        if i <= 9:
            sheetwrite.write(0, i, Excel_header[i], writestyle1)

        elif i > 9:
            sheetwrite.write(0, i, Excel_header[i], writestyle2)

        elif i > 20:
            sheetwrite.write(0, i, Excel_header[i], writestyle3)
    return(bookwrite, sheetwrite, sheetwrite2)

def Write_NETInfo(sheetwrite, n, CMXinfo):

    for i in range(0, Total_Col):
        sheetwrite.write(n, i, "", writestyle0)

    #NET part
    sheetwrite.write(n, 0, CMXinfo.Message_Name, writestyle0)
    sheetwrite.write(n, 1, CMXinfo.Signal_Name, writestyle0)

    i1 = Col_NET_n
    sheetwrite.write(n, i1, CMXinfo.Signal_Factor, writestyle0); i1 += 1
    sheetwrite.write(n, i1, CMXinfo.Signal_Offset, writestyle0); i1 += 1
    sheetwrite.write(n, i1, CMXinfo.Signal_PhyMin, writestyle0); i1 += 1
    sheetwrite.write(n, i1, CMXinfo.Signal_PhyMax, writestyle0); i1 += 1
    sheetwrite.write(n, i1, CMXinfo.Signal_ValueTab, writestyle0); i1 += 1
    sheetwrite.write(n, i1, 'InvalidVal: ' + CMXinfo.Signal_InvalidVal, writestyle0)

    #ASW part
    i2 = Col_ASW_n
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,2,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,3,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,4,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,5,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,6,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,7,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,8,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,9,FALSE)"), writestyle0);  i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,10,FALSE)"), writestyle0); i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,11,FALSE)"), writestyle0); i2 += 1
    sheetwrite.write(n, i2, xlwt.Formula("VLOOKUP(E1:E65536,Sheet2!A1:Z65536,12,FALSE)"), writestyle0)

class CMXinfo:
    def __init__ (self):
        self.Sender = ""
        self.Message_Name = ""
        self.Message_ID = 0
        self.Message_Sendtype = ""
        self.Message_Cycletime =  0
        self.Message_Length = 0
        self.Signal_Name = ""
        self.Signal_Byteorder = ""
        self.Signal_Startbyte = 0
        self.Signal_Startbit = 0
        self.Signal_Sendtype = ""
        self.Signal_Bitlength = 0
        self.Signal_Datatype = "unsigned"
        self.Signal_Factor = 0.0
        self.Signal_Offset = 0.0
        self.Signal_PhyMin = 0.0
        self.Signal_PhyMax = 0.0
        self.Signal_HexMin = 0x00
        self.Signal_HexMax = 0x00
        self.Signal_InitVal = 0x00
        self.Signal_InvalidVal = 0x00
        self.Signal_Unit = ""

        self.Signal_ValueTab = ""
        self.Message_ID_Hex = 0

    def Motorola_lsb2msb(self):
        row = int(self.Signal_Startbit/8)

        if self.Signal_Bitlength + self.Signal_Startbit <= (1+row)*8: #当前行
            msb = self.Signal_Startbit + self.Signal_Bitlength - 1
        else:
            shiftrow_l = int((self.Signal_Bitlength + self.Signal_Startbit - (1+row)*8)%8)
            if shiftrow_l == 0:
                shiftrow = int((self.Signal_Bitlength + self.Signal_Startbit - (1+row)*8)/8)
                msb = int((row-shiftrow + 1)*8 + shiftrow_l - 1)
            else:
                shiftrow = int((self.Signal_Bitlength + self.Signal_Startbit - (1+row)*8)/8 + 1)
                msb = int((row-shiftrow)*8 + shiftrow_l - 1)

        return msb

    def  Motorola_msb2lsb(self):
        row = int(self.Signal_Startbit/8)

        if  self.Signal_Startbit -self.Signal_Bitlength + 1 >= row*8: #当前行
            lsb = self.Signal_Startbit - self.Signal_Bitlength + 1
        else:
            shiftrow_l = int((self.Signal_Bitlength - (self.Signal_Startbit - row*8 +1))%8)
            if shiftrow_l == 0:
                shiftrow = int((self.Signal_Bitlength - (self.Signal_Startbit - row*8 +1))/8)
                lsb = int((row+shiftrow)*8 - shiftrow_l)
            else:
                shiftrow = int((self.Signal_Bitlength - (self.Signal_Startbit - row*8 +1))/8+1)
                lsb = int((row+shiftrow+1)*8 - shiftrow_l)

        return lsb

# function to store in csv file with pandas lib api
#def CsvStore(serdata):
#    dataframe = pandas.DataFrame(serdata)
#    dataframe.to_csv('NetRE_dbc_temp.csv', index=False, header=False)

class DbcParserClass:
    __messagePattern = r'(BO_)\s*(\d*)\s*(\w*)\s*:\s*(\d*)\s*(\w*)\s*\n'
    __cycletimePattern = r'(BA_)\s*("GenMsgCycleTime")\s*(BO_)\s*(\d*)\s*(\d*);\n'
    __signalPattern = r'(SG_)\s*(\w*\s*\w+)\s*:\s*(\d*)\s*\|\s*(\d*)\s*@\s*(\d*)\s*(\+|\-)\s*\(\s*(-?\d*\.*\d*)\s*,\s*(-?\d*\.*\d*)\s*\)\s*\[\s*(-?\d*\.*\d*)\s*\|\s*(-?\d*\.*\d*)\s*]\s*"(.*)"\s+([\w*,?]*)\s*\n+'
    __InvalidVluePattern = r'(BA_)\s*("GenSigInvalidValue")\s*(SG_)\s*(\d*)\s*(\w*)\s*"(.*)";\n'
    __InitVluePattern = r'(BA_)\s*("GenSigStartValue")\s*(SG_)\s*(\d*)\s*(\w*)\s*(.*);\n'
    __ValueTabPattern = r'(VAL_)\s*(\d*)\s*(\w*)\s*(.*)\s*;\n'
    __dbcFile = None
    __canMessages = None
    __canSignals = None


    # Constructor of
    def __init__(self, file, excel_file):
        # Initialize target file
        self.__dbcFile = file
        self.__excelFile = excel_file
        return

        # Return all message cycletime
    def getMessageCycletime(self):
        # If file is not yet parsed
        if self.__canMessages != None:
            try:
                # Open DBC file
                fileHandle = open(self.__dbcFile, 'r')

                # Iterate over all lines
                for line in fileHandle:
                    # If line matches a message description
                    if re.search(self.__cycletimePattern, line) != None:
                        # Get message description
                        tempDesc = re.findall(self.__cycletimePattern, line)
                        # Create a temporary message cycletime
                        tempMessageCycletime = {}
                        tempMessageCycletime['ID'] = int(tempDesc[0][3])
                        tempMessageCycletime['Cycletime'] = int(tempDesc[0][4])
                        # Add temporary message cycletime to CAN message list
                        for message in self.__canMessages:
                            if self.__canMessages[message]['ID'] == tempMessageCycletime['ID']:
                                self.__canMessages[message]['Cycletime'] = tempMessageCycletime['Cycletime']
            except Exception as ex:
                # Delete not finished list
                self.__canMessages = None
                print("Failed to parse DBC file (" + str(self.__dbcFile) + ").\n " + str(ex))
                return False
            finally:
                # Close file
                fileHandle.close()
        return self.__canMessages

    # Return a list of all messages
    def getMessageList(self):
        # If file is not yet parsed
        if self.__canMessages == None:
            try:
                # Create an empty message dictionary
                self.__canMessages = {}
                # Open DBC file
                fileHandle = open(self.__dbcFile, 'r')

                # Iterate over all lines
                for line in fileHandle:
                    # If line matches a message description
                    if re.search(self.__messagePattern, line) != None:
                        # Get message description
                        tempDesc = re.findall(self.__messagePattern, line)
                        # Create a temporary message
                        tempMessage = {}
                        tempMessage['ID'] = int(tempDesc[0][1])
                        # Set default value
                        tempMessage['Cycletime'] = 0
                        tempMessage['Name'] = tempDesc[0][2]
                        tempMessage['Size'] = int(tempDesc[0][3])
                        tempMessage['Transmitter'] = tempDesc[0][4]
                        tempMessage['Signals'] = {}
                        # Add temporary message to CAN message list
                        self.__canMessages[tempMessage['Name']] = tempMessage
                    # If line matches a signal description
                    elif re.search(self.__signalPattern, line) != None:
                        # Get signal description
                        tempDesc = re.findall(self.__signalPattern, line)
                        # Create a temporary signal
                        tempSignal = {}
                        tempSignal['Name'] = tempDesc[0][1]
                        tempSignal['Start-Bit'] = int(tempDesc[0][2])
                        tempSignal['Size'] = int(tempDesc[0][3])
                        tempSignal['Byte-Order'] = int(tempDesc[0][4])
                        tempSignal['Value-Type'] = tempDesc[0][5]
                        tempSignal['Factor'] = float(tempDesc[0][6])
                        tempSignal['Offset'] = float(tempDesc[0][7])
                        tempSignal['Minimum'] = float(tempDesc[0][8])
                        tempSignal['Maximum'] = float(tempDesc[0][9])
                        tempSignal['InvalidVlue'] = 'NA'
                        tempSignal['InitVlue'] = 'NA'
                        tempSignal['Unit'] = tempDesc[0][10]
                        tempSignal['ValueTab'] = 'NA'
                        #tempSignal['Receiver'] = tempDesc[12] # --> Receiver set not yet implemented
                        # Add temporary signal to current message of CAN message list
                        self.__canMessages[tempMessage['Name']]['Signals'][tempSignal['Name']] = tempSignal

                    elif re.search(self.__InvalidVluePattern, line) != None:
                        tempDesc1 = re.findall(self.__InvalidVluePattern, line)
                        message = 0
                        Isignal = 0
                        for message in self.__canMessages:
                            for Isignal in self.__canMessages[message]['Signals']:
                                if self.__canMessages[message]['Signals'][Isignal]['Name'] == (tempDesc1[0][4]):
                                    self.__canMessages[message]['Signals'][Isignal]['InvalidVlue'] = tempDesc1[0][5]

                    elif re.search(self.__InitVluePattern, line) != None:
                        tempDesc1 = re.findall(self.__InitVluePattern, line)
                        message = 0
                        Isignal = 0
                        for message in self.__canMessages:
                            for Isignal in self.__canMessages[message]['Signals']:
                                if self.__canMessages[message]['Signals'][Isignal]['Name'] == (tempDesc1[0][4]):
                                    self.__canMessages[message]['Signals'][Isignal]['InitVlue'] = tempDesc1[0][3]

                    elif re.search(self.__ValueTabPattern, line) != None:
                        tempDesc1 = re.findall(self.__ValueTabPattern, line)
                        message = 0
                        Isignal = 0
                        for message in self.__canMessages:
                            for Isignal in self.__canMessages[message]['Signals']:
                                if self.__canMessages[message]['Signals'][Isignal]['Name'] == (tempDesc1[0][2]):
                                    self.__canMessages[message]['Signals'][Isignal]['ValueTab'] = tempDesc1[0][3]
                #Add message cycle time to CAN message list
                self.getMessageCycletime()
            except Exception as ex:
                # Delete not finished list
                self.__canMessages = None
                print("Failed to parse DBC file (" + str(self.__dbcFile) + ").\n " + str(ex))
                return False
            finally:
                # Close file
                fileHandle.close()
        return self.__canMessages

    def NetRE_ParsedInformation(self):
        (bookwrite, sheetwrite) = ExcelWrite1(self.__excelFile)
        count = 0
        try:
            tempText = ""
            # If file is not yet parsed
            if self.__canMessages == None:
                self.getMessageList()
                logger.info("RE1_S1: getMessageList from Dbc file Parser according to pattern")
            # Print CAN message list
            for message in self.__canMessages:
                tempText = tempText + "Message: Name: " + str(message) + " ID: " + str(self.__canMessages[message]['ID']) + " Size: " + str(self.__canMessages[message]['Size']) + " Cycletime: " + str(self.__canMessages[message]['Cycletime']) + "\n"
                for signal in self.__canMessages[message]['Signals']:
                    tempText = tempText + "Signal:"
                    tempText = tempText + " Name:        " + str(signal)
                    tempText = tempText + " Start Bit:   " + str(self.__canMessages[message]['Signals'][signal]['Start-Bit'])
                    tempText = tempText + " Size:        " + str(self.__canMessages[message]['Signals'][signal]['Size'])
                    tempText = tempText + " Byte-Order:  " + str(self.__canMessages[message]['Signals'][signal]['Byte-Order'])
                    tempText = tempText + " Value-Type:  " + str(self.__canMessages[message]['Signals'][signal]['Value-Type'])
                    tempText = tempText + " Factor:      " + str(self.__canMessages[message]['Signals'][signal]['Factor'])
                    tempText = tempText + " Offset:      " + str(self.__canMessages[message]['Signals'][signal]['Offset'])
                    tempText = tempText + " Minimum:     " + str(self.__canMessages[message]['Signals'][signal]['Minimum'])
                    tempText = tempText + " Maximum:     " + str(self.__canMessages[message]['Signals'][signal]['Maximum'])
                    tempText = tempText + " InvalidVlue: " + str(self.__canMessages[message]['Signals'][signal]['InvalidVlue'])
                    tempText = tempText + " Unit:        " + str(self.__canMessages[message]['Signals'][signal]['Unit'])+ "\n"

                    #Write xls file
                    l_CMXinfo = CMXinfo()
                    l_CMXinfo. Message_Name       = str(message)
                    l_CMXinfo. Message_ID         = str(self.__canMessages[message]['ID'])
                    l_CMXinfo. Message_Cycletime  = str(self.__canMessages[message]['Cycletime'])
                    l_CMXinfo. Message_Length     = str(self.__canMessages[message]['Size'])
                    l_CMXinfo. Sender = str(self.__canMessages[message]['Transmitter'])

                    l_CMXinfo. Signal_Name        = str(signal)
                    l_CMXinfo. Signal_Byteorder   = int(self.__canMessages[message]['Signals'][signal]['Byte-Order'])
                    l_CMXinfo. Signal_Startbit    = int(self.__canMessages[message]['Signals'][signal]['Start-Bit'])
                    l_CMXinfo. Signal_Bitlength   = int(self.__canMessages[message]['Signals'][signal]['Size'])
                    l_CMXinfo. Signal_Datatype    = str(self.__canMessages[message]['Signals'][signal]['Value-Type'])
                    l_CMXinfo. Signal_Factor      = float(self.__canMessages[message]['Signals'][signal]['Factor'])
                    l_CMXinfo. Signal_Offset      = float(self.__canMessages[message]['Signals'][signal]['Offset'])
                    l_CMXinfo. Signal_PhyMin      = float(self.__canMessages[message]['Signals'][signal]['Minimum'])
                    l_CMXinfo. Signal_PhyMax      = float(self.__canMessages[message]['Signals'][signal]['Maximum'])
                    l_CMXinfo. Signal_InvalidVal  = str(self.__canMessages[message]['Signals'][signal]['InvalidVlue'])
                    l_CMXinfo. Signal_InitVal     = str(self.__canMessages[message]['Signals'][signal]['InitVlue'])
                    l_CMXinfo. Signal_ValueTab     = str(self.__canMessages[message]['Signals'][signal]['ValueTab'])

                    if l_CMXinfo. Signal_InvalidVal.isdigit():
                        l_CMXinfo. Signal_InvalidVal = hex(int(l_CMXinfo. Signal_InvalidVal))
                    if l_CMXinfo. Signal_InitVal.isdigit():
                        l_CMXinfo. Signal_InitVal = hex(int(l_CMXinfo. Signal_InitVal))
                    if l_CMXinfo. Message_ID.isdigit():
                        l_CMXinfo. Message_ID_Hex     = hex(int(l_CMXinfo. Message_ID)).upper()
                    if l_CMXinfo. Signal_Byteorder == 0:
                        l_CMXinfo. Signal_Byteorder = 'MotorolaLSB'#CMX:LSB, DBC:MSB
                        l_CMXinfo. Signal_Startbit = l_CMXinfo.Motorola_msb2lsb()
                    else:
                        l_CMXinfo. Signal_Byteorder = 'Intel'#CMX:LSB, DBC:MSB

                    count = count + 1

                    count = Write_NETInfo1(sheetwrite, count, l_CMXinfo)


            Adjust_Style(sheetwrite)
            #split temp text by "\n"
            #SplittempText = tempText.split("\n")
            #stort to csv file
            logger.info("RE1_S2: Store final NET information to file ")
            #CsvStore(SplittempText)
            return True
        except Exception as ex:
            logger.info("Failed to print DBC information of (" + str(self.__dbcFile) + ").\n " + str(ex))
            return False
        finally:
                # Save Close file
                bookwrite.save(self.__excelFile)
